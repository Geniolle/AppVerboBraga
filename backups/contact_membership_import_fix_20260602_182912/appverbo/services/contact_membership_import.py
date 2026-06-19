from __future__ import annotations

import csv
import io
import unicodedata
from typing import Any

from openpyxl import load_workbook


# ###################################################################################
# (1) NORMALIZACAO E CABECALHOS
# ###################################################################################

HEADER_ALIASES_V1: dict[str, tuple[str, ...]] = {
    "name": ("nome", "name"),
    "phone": ("telefone", "telemovel", "telemovel", "celular", "phone"),
    "email": ("email", "e-mail", "mail"),
}


def _normalize_lookup_text_v1(raw_value: Any) -> str:
    normalized = (
        unicodedata.normalize("NFKD", str(raw_value or ""))
        .encode("ascii", "ignore")
        .decode("ascii")
        .strip()
        .lower()
    )
    return " ".join(normalized.split())


def _normalize_cell_text_v1(raw_value: Any) -> str:
    if raw_value is None:
        return ""
    if isinstance(raw_value, bool):
        return "1" if raw_value else "0"
    if isinstance(raw_value, int):
        return str(raw_value).strip()
    if isinstance(raw_value, float):
        if raw_value.is_integer():
            return str(int(raw_value)).strip()
        return str(raw_value).strip()
    return str(raw_value).strip()


def _resolve_header_indexes_v1(headers: list[str]) -> dict[str, int]:
    normalized_headers = [_normalize_lookup_text_v1(header) for header in headers]
    indexes: dict[str, int] = {}

    for role_key, aliases in HEADER_ALIASES_V1.items():
        for index, normalized_header in enumerate(normalized_headers):
            if normalized_header in aliases:
                indexes[role_key] = index
                break

    missing_roles = [role_key for role_key in ("name", "phone", "email") if role_key not in indexes]
    if missing_roles:
        raise ValueError(
            "O ficheiro deve conter os cabeçalhos Nome, Telefone e Email."
        )

    return indexes


# ###################################################################################
# (2) LEITURA DE CSV E XLSX
# ###################################################################################

def _read_csv_rows_v1(raw_bytes: bytes) -> list[list[str]]:
    decoded_text = ""
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            decoded_text = raw_bytes.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if not decoded_text:
        raise ValueError("Nao foi possivel ler o ficheiro CSV.")

    sample = decoded_text[:2048]
    delimiter = ";"
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
        delimiter = str(getattr(dialect, "delimiter", ";") or ";")
    except csv.Error:
        delimiter = ";"

    reader = csv.reader(io.StringIO(decoded_text), delimiter=delimiter)
    return [
        [_normalize_cell_text_v1(cell) for cell in row]
        for row in reader
    ]


def _read_xlsx_rows_v1(raw_bytes: bytes) -> list[list[str]]:
    workbook = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    try:
        worksheet = workbook.worksheets[0] if workbook.worksheets else workbook.active
        if worksheet is None:
            return []
        return [
            [_normalize_cell_text_v1(cell) for cell in row]
            for row in worksheet.iter_rows(values_only=True)
        ]
    finally:
        workbook.close()


# ###################################################################################
# (3) PARSER PUBLICO
# ###################################################################################

async def parse_contact_membership_import_file_v1(upload_file: Any) -> dict[str, Any]:
    filename = str(getattr(upload_file, "filename", "") or "").strip()
    if not filename:
        raise ValueError("Selecione um ficheiro CSV ou XLSX para importar.")

    extension = filename.rsplit(".", 1)[-1].strip().lower() if "." in filename else ""
    if extension not in {"csv", "xlsx"}:
        raise ValueError("Formato invalido. Use um ficheiro CSV ou XLSX.")

    raw_bytes = await upload_file.read()
    if not raw_bytes:
        raise ValueError("O ficheiro enviado esta vazio.")

    if extension == "csv":
        raw_rows = _read_csv_rows_v1(raw_bytes)
    else:
        raw_rows = _read_xlsx_rows_v1(raw_bytes)

    if not raw_rows:
        raise ValueError("O ficheiro nao contem dados para importar.")

    headers = [_normalize_cell_text_v1(cell) for cell in (raw_rows[0] or [])]
    header_indexes = _resolve_header_indexes_v1(headers)

    imported_rows: list[dict[str, str]] = []
    skipped_rows = 0

    for raw_row in raw_rows[1:]:
        normalized_row = [_normalize_cell_text_v1(cell) for cell in raw_row]

        def _get_value(role_key: str) -> str:
            target_index = header_indexes[role_key]
            if target_index >= len(normalized_row):
                return ""
            return normalized_row[target_index]

        clean_name = _get_value("name")
        clean_phone = _get_value("phone")
        clean_email = _get_value("email")

        if not clean_name and not clean_phone and not clean_email:
            skipped_rows += 1
            continue

        if not clean_name or not clean_phone or not clean_email:
            skipped_rows += 1
            continue

        imported_rows.append(
            {
                "name": clean_name,
                "phone": clean_phone,
                "email": clean_email,
            }
        )

    return {
        "filename": filename,
        "rows": imported_rows,
        "skipped_rows": skipped_rows,
    }
